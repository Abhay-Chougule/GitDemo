import openpyxl


class HomePageData:

    test_homepage_data = [{"Name": "Abhi", "Email": "Kore", "Gender": "Male"}, {"Name": "Jeni", "Email": "Doe", "Gender": "Female"}]

    @staticmethod
    def getTestData(testCaseName):
        filepath = "C:\\Users\\abhay.chougule\\PycharmProjects\\PythonSelfFramework\\TestData\\HomePageTestData.xlsx"
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook["Data1"]

        dicti = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testCaseName:
                for j in range(2, sheet.max_column + 1):
                    # dict[sheet.cell(row=1, column=j).value] = (sheet.cell(row=i + 1, column=j).value)
                    dicti.update({sheet.cell(row=1, column=j).value: sheet.cell(row=i, column=j).value})
                break
        return[dicti]


import openpyxl

filepath = "C:\\Users\\abhay.chougule\\PycharmProjects\\PythonSelfFramework\\TestData\\HomePageTestData.xlsx"
workbook = openpyxl.load_workbook(filepath)
sheet = workbook["Data1"]
cell1 = sheet.cell(row=1, column=2)
print(cell1.value)

sheet.cell(row=2, column=2).value = "Abhay"

workbook.save(filepath)

NoRow = sheet.max_row
NoColumn = sheet.max_column
print(NoRow)
print(NoColumn)

# for i in range(1, NoRow+1):
#     for j in range(1, NoColumn+1):
#         print(sheet.cell(row=i, column=j).value, end=", ")
#     print()

dict = {}
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "TestCase1":
        for j in range(1, sheet.max_column+1):
            # dict[sheet.cell(row=1, column=j).value] = (sheet.cell(row=i + 1, column=j).value)
            dict.update({sheet.cell(row=1, column=j).value : sheet.cell(row=i, column=j).value})
        break
print(dict)
